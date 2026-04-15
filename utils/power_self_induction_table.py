# -*- coding: utf-8 -*-
from __future__ import annotations
import numpy as np
import pandas as pd
from pathlib import Path

# ================== НАСТРОЙКИ ==================
MODEL_CSV   = r"D:\PROJECTs\magnet\harvester\graphs\experiment_2025-10-22_11-45-42\2025-10-22_11-45-42_results_model_raw.csv"
OUT_DIR     = Path("graphs/comparison_results"); OUT_DIR.mkdir(parents=True, exist_ok=True)

R_LOAD_OHM  = 1.9     # Ом
RANGE_SELECT: tuple[float, float] | None = (7.0, 12.0)   # None чтобы отключить
MAKE_PLOT   = True            # Сохранить график мощностей с/без самоиндукции

# печать в консоль — научная нотация, чтобы видеть малые отличия
pd.set_option("display.float_format", "{:.10e}".format)


# =============== УТИЛИТЫ ===============
def read_model_csv(csv_path: str) -> pd.DataFrame:
    """Ожидает колонки: freq_Hz, model_rms_V, model_self_rms_V (регистр не критичен)."""
    df = pd.read_csv(csv_path)
    cols = {c.lower(): c for c in df.columns}

    def pick(*keys):
        for k in keys:
            if k in cols: return cols[k]
        return None

    fcol   = pick("freq_hz") \
             or next((c for c in df.columns if "freq" in c.lower()), None)
    vcol   = pick("model_rms_v")
    vself  = pick("model_self_rms_v")

    if fcol is None:
        raise ValueError("В файле не найдена колонка частоты (freq*).")
    if vcol is None:
        raise ValueError("В файле не найдена колонка model_rms_V (индукционный ЭДС).")
    if vself is None:
        raise ValueError("В файле не найдена колонка model_self_rms_V (самоиндукционный ЭДС).")

    df = df[[fcol, vcol, vself]].rename(columns={
        fcol:  "freq_Hz",
        vcol:  "model_rms_V",
        vself: "model_self_rms_V",
    }).dropna().sort_values("freq_Hz").reset_index(drop=True)

    return df


def apply_range(df: pd.DataFrame, rng: tuple[float, float] | None) -> pd.DataFrame:
    if not rng:
        return df
    lo, hi = rng
    return df[(df["freq_Hz"] >= lo) & (df["freq_Hz"] <= hi)].reset_index(drop=True)


def fwhm_interpolated(freq: np.ndarray, y: np.ndarray) -> float:
    """
    Ширина на полувысоте (FWHM) с линейной интерполяцией краёв.
    Возвращает NaN, если невозможно корректно оценить.
    """
    y = np.asarray(y, float)
    if y.size < 3 or not np.isfinite(y).any():
        return np.nan

    idx_max = int(np.nanargmax(y))
    y_max = y[idx_max]
    if not np.isfinite(y_max) or y_max <= 0:
        return np.nan

    half = 0.5 * y_max

    # влево от пика
    left_x = np.nan
    i = idx_max
    while i > 0 and y[i] >= half:
        i -= 1
    if i < idx_max and np.isfinite(y[i]) and np.isfinite(y[i+1]) and y[i] <= half <= y[i+1]:
        x0, x1 = freq[i], freq[i+1]
        y0, y1 = y[i], y[i+1]
        left_x = x0 + (half - y0) * (x1 - x0) / (y1 - y0 + 1e-12)

    # вправо от пика
    right_x = np.nan
    i = idx_max
    n = y.size
    while i < n - 1 and y[i] >= half:
        i += 1
    if i > idx_max and np.isfinite(y[i-1]) and np.isfinite(y[i]) and y[i-1] >= half >= y[i]:
        x0, x1 = freq[i-1], freq[i]
        y0, y1 = y[i-1], y[i]
        right_x = x0 + (half - y0) * (x1 - x0) / (y1 - y0 + 1e-12)

    if np.isfinite(left_x) and np.isfinite(right_x) and right_x >= left_x:
        return float(right_x - left_x)
    return np.nan


def build_table6_from_rows(freq: np.ndarray, V_wo: np.ndarray, V_with: np.ndarray, r_load_ohm: float) -> pd.DataFrame:
    """Формирует Table 6 по двум рядам RMS-напряжений."""
    freq = np.asarray(freq, float)
    V_wo = np.asarray(V_wo, float)
    V_with = np.asarray(V_with, float)

    # мощности (мВт)
    P_wo   = (V_wo**2)   / r_load_ohm * 1000.0
    P_with = (V_with**2) / r_load_ohm * 1000.0

    # метрики
    Pavg_wo, Pavg_with = float(np.mean(P_wo)), float(np.mean(P_with))
    Vp_wo, Vp_with     = float(np.max(V_wo)), float(np.max(V_with))

    BW_wo   = fwhm_interpolated(freq, P_wo)
    BW_with = fwhm_interpolated(freq, P_with)

    f_res_wo   = float(freq[int(np.argmax(P_wo))])
    f_res_with = float(freq[int(np.argmax(P_with))])

    def rel_signed(with_v, wo_v):
        return float(abs(with_v - wo_v) / wo_v * 100.0) if wo_v != 0 else np.nan

    table = pd.DataFrame([
        ["Average Load Power, mW",                 Pavg_wo, Pavg_with, rel_signed(Pavg_with, Pavg_wo)],
        ["Amplitude of Output Voltage (Vp), V",    Vp_wo,   Vp_with,   rel_signed(Vp_with, Vp_wo)],
        ["Bandwidth of Resonance Peak by Power, Hz", BW_wo, BW_with,   rel_signed(BW_with, BW_wo)],
        ["Maximum Resonance Frequency, Hz",        f_res_wo, f_res_with, rel_signed(f_res_with, f_res_wo)],
    ], columns=["Parameter", "Without Self-Induction", "With Self-Induction", "Relative change (%)"])

    return table


def save_table_excel(table: pd.DataFrame, xlsx_path: Path) -> None:
    """
    Сохраняет xlsx с корректными числовыми форматами и
    дивергирующей подсветкой для столбца 'Relative change (%)' вокруг 0.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.formatting.rule import ColorScaleRule

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Table6", index=False)

    wb = load_workbook(xlsx_path)
    ws = wb["Table6"]

    # шапка
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # авто-ширина и форматы
    col_map = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    num_fmt_sci = "0.0000000000E+00"
    num_fmt_std = "0.0000000000"
    pct_fmt     = "0.0000000000"

    # применяем форматы
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # столбец 2..4 — числовые
        for col_idx in (2, 3, 4):
            cell = row[col_idx - 1]
            # относительные изменения — в %, но это уже числа (не доли)
            if col_idx == 4:
                cell.number_format = pct_fmt
            else:
                # выбираем научную нотацию (чтобы не «съедало» малые)
                cell.number_format = num_fmt_sci

    # авто-ширина
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    # подсветка для 'Relative change (%)' вокруг 0 (дивергирующая шкала)
    idx = col_map.get("Relative change (%)")
    if idx is not None:
        # найдём симметричные границы по модулю
        vals = []
        for r in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            v = r[0].value
            if isinstance(v, (int, float)):
                vals.append(float(v))
        if vals:
            bound = max(abs(min(vals)), abs(max(vals)))
            c1 = ws.cell(row=2, column=idx).coordinate
            c2 = ws.cell(row=ws.max_row, column=idx).coordinate
            rule = ColorScaleRule(
                start_type='num', start_value=-bound, start_color='63BE7B',
                mid_type='num',   mid_value=0,       mid_color='FFEB84',
                end_type='num',   end_value=+bound,  end_color='F8696B'
            )
            ws.conditional_formatting.add(f"{c1}:{c2}", rule)

    wb.save(xlsx_path)


def plot_powers(freq: np.ndarray, P_wo_mW: np.ndarray, P_with_mW: np.ndarray, out_path: Path) -> None:
    """График мощностей (мВт) без/с самоиндукцией. Красная — с самоиндукцией."""
    import matplotlib.pyplot as plt

    FIGSIZE_INCH = (6, 3)    # 2:1
    SAVE_DPI = 1200
    FONT_FAMILY = "Arial"
    FONT_SIZE = 12
    XTICK_STEP = 0.5

    plt.rcParams.update({
        "font.family": FONT_FAMILY,
        "font.size": FONT_SIZE,
        "savefig.dpi": SAVE_DPI,
        "axes.labelsize": FONT_SIZE,
        "axes.titlesize": FONT_SIZE,
        "legend.fontsize": FONT_SIZE,
        "xtick.labelsize": FONT_SIZE,
        "ytick.labelsize": FONT_SIZE,
    })

    plt.figure(figsize=FIGSIZE_INCH)
    ax = plt.gca()
    ax.plot(freq, P_wo_mW, "-", linewidth=2, label="Power w/o Self-Induction")      # по умолчанию — чёрная
    ax.plot(freq, P_with_mW, "-", linewidth=2, color="red", label="Power with Self-Induction")  # красная

    ax.set_xlabel("Frequency (Hz)")
    ax.set_ylabel("Power (mW)")
    ax.grid(True)

    xmin, xmax = np.floor(freq.min()), np.ceil(freq.max())
    ax.set_xlim(xmin, xmax)
    ax.set_xticks(np.arange(xmin, xmax + XTICK_STEP, XTICK_STEP))
    ax.legend()
    plt.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(out_path, bbox_inches="tight")
    plt.close()


# =============== ОСНОВНОЙ ХОД ===============
def main():
    df = read_model_csv(MODEL_CSV)
    df = apply_range(df, RANGE_SELECT)

    if df.empty:
        raise RuntimeError("После фильтрации по диапазону данных не осталось.")

    f     = df["freq_Hz"].to_numpy(dtype=float)
    Vind  = df["model_rms_V"].to_numpy(dtype=float)          # индукционный ЭДС (без L)
    Vself = df["model_self_rms_V"].to_numpy(dtype=float)     # ЭДС самоиндукции

    # Режим без самоиндукции (L=0): просто индукционный ЭДС
    V_wo = Vind

    # Режим с самоиндукцией: самоиндукция противодействует — вычитаем (по модулю RMS)
    # ВНИМАНИЕ: строго говоря, RMS(|Vind - Vself|) корректен при фазовом совпадении;
    # для твоих данных эффект малый — аппроксимация уместна.
    V_with = np.abs(Vind - Vself)

    # Мощности (мВт) — для графика и метрик
    P_wo_mW   = (V_wo**2)   / R_LOAD_OHM * 1000.0
    P_with_mW = (V_with**2) / R_LOAD_OHM * 1000.0

    # Таблица 6
    table6 = build_table6_from_rows(f, V_wo, V_with, R_LOAD_OHM)

    # Печать (научная нотация)
    print("\n=== Table 6 (Self-Induction Effect) ===")
    print(table6)

    # Сохранить Excel с форматами (ничего не «съест»)
    xlsx_path = OUT_DIR / "table6_self_induction.xlsx"
    save_table_excel(table6, xlsx_path)
    print(f"[OK] Excel saved -> {xlsx_path}")

    # График (опционально)
    if MAKE_PLOT:
        plot_path = OUT_DIR / "power_with_vs_without_self.png"
        plot_powers(f, P_wo_mW, P_with_mW, plot_path)
        print(f"[OK] Plot saved -> {plot_path}")


if __name__ == "__main__":
    main()
