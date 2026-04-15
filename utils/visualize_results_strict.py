# -*- coding: utf-8 -*-
import numpy as np
import pandas as pd
from pathlib import Path
import matplotlib.pyplot as plt

# === ПУТИ ===
MODEL_CSV = r"D:\PROJECTs\magnet\harvester\graphs\experiment_2025-10-22_11-45-42\2025-10-22_11-45-42_results_model_raw.csv"
MODEL_CSV = r"D:\PROJECTs\magnet\harvester\graphs\experiment_2025-10-16_13-35-09\2025-10-16_13-35-09_results_model_raw.csv"
EXP_CSV   = r"experiments\harvester_80mm\rms_compact_freq_rms.csv"
MODEL_CSV = r"D:\PROJECTs\magnet\harvester\graphs\experiment_2025-10-22_11-45-42\2025-10-22_11-45-42_results_model_raw.csv"

OUT_DIR = Path("graphs/comparison_results"); OUT_DIR.mkdir(parents=True, exist_ok=True)

# === ПАРАМЕТРЫ ===
RANGE_SELECT = (7.0, 12.0)   # None чтобы отключить
MODEL_SCALE = 1.0
EXP_SCALE   = 1.0
AUTO_SCALE  = None           # "model_to_exp" | "exp_to_model" | None
R_LOAD_OHM  = 1.9      # для RAW мощности (мВт)
OFFSET_MODEL = 0.0000275     # оффсет из твоего примера; 0.0 если не нужен
EPS_FRAC = 0.05

from matplotlib.ticker import ScalarFormatter

def plot_self_induction_power(
    model_csv: str,
    out_dir: Path,
    r_load_ohm: float = 1.90,
    range_select: tuple | None = None,
    xtick_step: float = 0.5,
    figsize_inch: tuple = (6, 3),
    filename: str = "self_induction_power.png",
) -> Path:

    raw = pd.read_csv(model_csv)
    cols_l = {c.lower(): c for c in raw.columns}

    # freq col
    if "freq_hz" in cols_l:
        fcol = cols_l["freq_hz"]
    else:
        fcol = next((c for c in raw.columns if "freq" in c.lower()), None)
        if fcol is None:
            raise ValueError("В файле нет колонки частоты (freq*).")

    # self rms col
    if "model_self_rms_v" in cols_l:
        scol = cols_l["model_self_rms_v"]
    else:
        candidates = [c for c in raw.columns if all(k in c.lower() for k in ("self", "rms", "v"))]
        scol = candidates[0] if candidates else None
    if scol is None:
        raise ValueError("Не найдена колонка RMS самоиндукции (model_self_rms_V).")

    df = raw[[fcol, scol]].rename(columns={fcol: "freq_Hz", scol: "self_rms_V"})
    df = df.dropna().sort_values("freq_Hz").reset_index(drop=True)

    if range_select:
        lo_r, hi_r = range_select
        df = df[(df["freq_Hz"] >= lo_r) & (df["freq_Hz"] <= hi_r)]

    if df.empty:
        raise ValueError("После фильтрации данных не осталось точек для построения графика.")

    freq = df["freq_Hz"].to_numpy(float)
    v    = df["self_rms_V"].to_numpy(float)
    p_mW = (v * v) / float(r_load_ohm) * 1000.0

    plt.figure(figsize=figsize_inch)
    ax = plt.gca()


    ax.plot(freq, p_mW, "o-", color="red", linewidth=2, label="Self-induction Power")
    ax.set_xlabel("Frequency (Hz)")
    ax.set_ylabel("Power (mW)")
    ax.grid(True)

    # ---- ВКЛЮЧАЕМ SCI-ФОРМАТ ДЛЯ Y (без этого offset пустой) ----
    fmt = ScalarFormatter(useMathText=True)
    fmt.set_scientific(True)
    fmt.set_powerlimits((0, 0))     # всегда sci
    ax.yaxis.set_major_formatter(fmt)

    # чтобы Matplotlib пересчитал подписи/offset:
    ax.relim()
    ax.autoscale_view()
    ax.figure.canvas.draw()
 

    offset = ax.yaxis.get_offset_text()
    print(f"[INFO] Offset text before: {offset.get_text()}")  # для отладки
    # берём реальную степень автоматически
    exp = ax.yaxis.get_major_formatter().orderOfMagnitude

    offset.set_text(rf"$1\times10^{{{exp}}}$")
    offset.set_fontsize(plt.rcParams.get("font.size", 12))

    # ---- НАСИЛЬНО ПИШЕМ "1×10^n" ----
                                                # # orderOfMagnitude берём из форматтера (реальная степень, не хардкод)
                                                # exp = ax.yaxis.get_major_formatter().orderOfMagnitude
                                                # ax.yaxis.get_offset_text().set_text(rf"$1\times10^{{{exp}}}$")
                                                # ax.yaxis.get_offset_text().set_fontsize(plt.rcParams.get("font.size", 12))

    # X ticks
    xmin, xmax = np.floor(freq.min()), np.ceil(freq.max())
    ax.set_xlim(xmin, xmax)
    ax.set_xticks(np.arange(xmin, xmax + xtick_step, xtick_step))

    ax.legend()
    # plt.tight_layout()

    out_path = Path(out_dir) / filename
    out_path.parent.mkdir(parents=True, exist_ok=True)
    print(f"[INFO] Offset text after: {offset.get_text()}")  # для отладки

    plt.savefig(out_path, dpi=plt.rcParams.get("savefig.dpi", 1200), bbox_inches="tight")
    plt.close()

    print(f"[OK] Self-induction power plot saved: {out_path}")
    return out_path

# === УТИЛИТЫ ===
def load_csv_generic(path: str, scale: float) -> pd.DataFrame:
    df = pd.read_csv(path)
    low = [c.lower() for c in df.columns]
    if "model_rms_v" in low:
        freq_col = next(c for c in df.columns if "freq" in c.lower())
        rms_col  = next(c for c in df.columns if "rms"  in c.lower())
        df = df.rename(columns={freq_col: "freq_Hz", rms_col: "rms_V"})
    elif "rms_emf_mv" in low:
        freq_col = next(c for c in df.columns if "freq" in c.lower())
        rms_col  = next(c for c in df.columns if "rms"  in c.lower())
        df = df.rename(columns={freq_col: "freq_Hz"})
        df["rms_V"] = pd.to_numeric(df[rms_col], errors="coerce") * 1e-3
    elif "rms (v)" in low:
        df = df.rename(columns={"Frequency (Hz)": "freq_Hz", "RMS (V)": "rms_V"})
    else:
        freq_col = next((c for c in df.columns if "freq" in c.lower()), None)
        rms_col  = next((c for c in df.columns if "rms"  in c.lower()), None)
        if not freq_col or not rms_col:
            raise ValueError(f"Не распознан формат CSV: {path}")
        scale_guess = 1e-3 if "mv" in (rms_col or "").lower() else 1.0
        df = df.rename(columns={freq_col: "freq_Hz"})
        df["rms_V"] = pd.to_numeric(df[rms_col], errors="coerce") * scale_guess
    df = df[["freq_Hz", "rms_V"]].dropna().sort_values("freq_Hz").reset_index(drop=True)
    df["rms_V"] = df["rms_V"] * float(scale)
    return df

def minmax_norm(y: np.ndarray) -> np.ndarray:
    y = np.asarray(y, float)
    if y.size == 0:
        return y
    y_min, y_max = np.min(y), np.max(y)
    return np.zeros_like(y) if np.isclose(y_min, y_max) else (y - y_min) / (y_max - y_min)

def fit_scale(y_ref: np.ndarray, y_to_scale: np.ndarray) -> float:
    num = np.sum(y_ref * y_to_scale)
    den = np.sum(y_to_scale**2)
    return float(num / den) if den > 0 else 1.0

# === ЗАГРУЗКА И ПОДГОТОВКА ===
model_df = load_csv_generic(MODEL_CSV, MODEL_SCALE)
exp_df   = load_csv_generic(EXP_CSV,   EXP_SCALE)

if RANGE_SELECT:
    lo_r, hi_r = RANGE_SELECT
    model_df = model_df[(model_df["freq_Hz"] >= lo_r) & (model_df["freq_Hz"] <= hi_r)]
    exp_df   = exp_df[(exp_df["freq_Hz"] >= lo_r) & (exp_df["freq_Hz"]   <= hi_r)]

mf = model_df["freq_Hz"].to_numpy()
ef = exp_df["freq_Hz"].to_numpy()
ev = exp_df["rms_V"].to_numpy()

lo = max(np.min(mf), np.min(ef)); hi = min(np.max(mf), np.max(ef))
mask_overlap = (mf >= lo) & (mf <= hi)

exp_interp_V = np.full_like(mf, np.nan, dtype=float)
exp_interp_V[mask_overlap] = np.interp(mf[mask_overlap], ef, ev)

model_V = model_df["rms_V"].to_numpy() - OFFSET_MODEL
# ====== АВТО-ПОДГОНКА АМПЛИТУДЫ ======
# Режимы:
#   None            – без подгонки
#   "model_to_exp"  – старый: s*model ≈ exp (через ноль)
#   "exp_to_model"  – старый: s*exp   ≈ model (через ноль)
#   "affine"        – НОВОЕ: e ≈ a*m + b  (веса опц.)
#   "median"        – НОВОЕ: робастный масштаб по медиане отношений
#   "wls"           – НОВОЕ: взвешенная МНК через ноль (весим пик эксперимента)
AUTO_SCALE_MODE = "affine"   # поменяй при необходимости: "affine" | "median" | "wls" | None

applied_scale  = 1.0
applied_offset = 0.0

def fit_affine(y_ref: np.ndarray, y_pred: np.ndarray, mask: np.ndarray, w: np.ndarray | None = None):
    """
    Находит a, b из y_ref ≈ a * y_pred + b на пересечении mask.
    Если w задан — WLS (взвешенные наим. квадраты), иначе обычная МНК.
    """
    x = y_pred[mask].astype(float)
    y = y_ref[mask].astype(float)
    if x.size == 0:
        return 1.0, 0.0
    if w is not None:
        w = np.sqrt(np.clip(w[mask].astype(float), 0.0, np.inf))
        X = np.vstack([x, np.ones_like(x)]).T
        Xw = X * w[:, None]
        yw = y * w
        a, b = np.linalg.lstsq(Xw, yw, rcond=None)[0]
    else:
        X = np.vstack([x, np.ones_like(x)]).T
        a, b = np.linalg.lstsq(X, y, rcond=None)[0]
    return float(a), float(b)

if AUTO_SCALE_MODE is None:
    pass

elif AUTO_SCALE_MODE == "affine":
    # Сильнее «весим» область пика эксперимента
    eN = minmax_norm(exp_interp_V)          # 0..1 на всей сетке mf
    weights = (eN ** 3) * mask_overlap      # p=3; можно 2..4
    a, b = fit_affine(exp_interp_V, model_V, mask_overlap, w=weights)
    model_V = a * model_V + b
    applied_scale, applied_offset = a, b
    print(f"[AUTO_SCALE affine] a={a:.6f}, b={b:.6e}  ->  model := a*model + b")

elif AUTO_SCALE_MODE == "median":
    # Робастный одномерный масштаб: s = median(exp/model)
    ratios = exp_interp_V[mask_overlap] / np.maximum(model_V[mask_overlap], 1e-12)
    ratios = ratios[np.isfinite(ratios) & (ratios > 0)]
    s = np.median(ratios) if ratios.size else 1.0
    model_V *= s
    applied_scale = s
    print(f"[AUTO_SCALE median] s={s:.6f}  ->  model := s*model")

elif AUTO_SCALE_MODE == "wls":
    # Взвешенная МНК через ноль: e ≈ s*m, вес = (norm(exp))^p
    eN = minmax_norm(exp_interp_V)
    w = (eN ** 3) * mask_overlap
    m = model_V
    e = exp_interp_V
    num = np.sum(w * e * m)
    den = np.sum(w * m * m)
    s = float(num / den) if den > 0 else 1.0
    model_V *= s
    applied_scale = s
    print(f"[AUTO_SCALE wls] s={s:.6f}  ->  model := s*model")

else:
    # Поддержка твоих старых режимов на случай, если ты их используешь параллельно
    if AUTO_SCALE in ("model_to_exp", "exp_to_model"):
        if AUTO_SCALE == "model_to_exp":
            s = fit_scale(exp_interp_V[mask_overlap], model_V[mask_overlap])
            model_V *= s
            applied_scale = s
            print(f"[AUTO_SCALE legacy model_to_exp] s={s:.6f}")
        else:
            s = fit_scale(model_V[mask_overlap], exp_interp_V[mask_overlap])
            exp_interp_V[mask_overlap] *= s
            applied_scale = s
            print(f"[AUTO_SCALE legacy exp_to_model] s={s:.6f}")


# === RAW ТАБЛИЦА ===
freq_overlap = mf[mask_overlap].astype(float)
v_model_raw  = model_V[mask_overlap].astype(float)
v_exp_raw    = exp_interp_V[mask_overlap].astype(float)

p_model_mW = (v_model_raw**2) / R_LOAD_OHM * 1000.0
p_exp_mW   = (v_exp_raw**2)  / R_LOAD_OHM * 1000.0

eps = np.finfo(float).eps
rel_err_rms_raw   = (np.abs(v_model_raw - v_exp_raw)) / np.where(np.abs(v_exp_raw) > 0, v_exp_raw, eps) * 100.0
rel_err_power_raw = (np.abs(p_model_mW - p_exp_mW))   / np.where(np.abs(p_exp_mW) > 0, p_exp_mW, eps) * 100.0

df_raw = pd.DataFrame({
    "Frequency, Hz": freq_overlap,
    "RMS Voltage Model, V": v_model_raw,
    "RMS Voltage Experiment, V": v_exp_raw,
    "Output Power Model, mW": p_model_mW,
    "Output Power Experiment, mW": p_exp_mW,
    "Relative Error RMS, %": rel_err_rms_raw,
    "Relative Error Power, %": rel_err_power_raw,
})

# === NORMALIZED ТАБЛИЦА (ошибки ПОСЛЕ нормализаций) ===
mN_full = minmax_norm(model_V)
eN_full = minmax_norm(exp_interp_V)

v_model_norm = mN_full[mask_overlap].astype(float)
v_exp_norm   = eN_full[mask_overlap].astype(float)

p_model_norm = v_model_norm**2
p_exp_norm   = v_exp_norm**2

rel_err_rms_norm   = (v_model_norm - v_exp_norm) / np.where(np.abs(v_exp_norm) > 0, v_exp_norm, eps) * 100.0
rel_err_power_norm = (p_model_norm - p_exp_norm) / np.where(np.abs(p_exp_norm) > 0, p_exp_norm, eps) * 100.0

df_norm = pd.DataFrame({
    "Frequency, Hz": freq_overlap,
    "RMS Voltage Model, V (raw)": v_model_raw,
    "RMS Voltage Experiment, V (raw)": v_exp_raw,
    "RMS Voltage Model (normalized)": v_model_norm,
    "RMS Voltage Experiment (normalized)": v_exp_norm,
    "Power Model (normalized, V^2)": p_model_norm,
    "Power Experiment (normalized, V^2)": p_exp_norm,
    "Relative Error RMS, % (normalized)": rel_err_rms_norm,
    "Relative Error Power, % (normalized)": rel_err_power_norm,
})

# === СОХРАНЕНИЕ CSV ===
csv_raw   = OUT_DIR / "pointwise_comparison_raw.csv"
csv_norm  = OUT_DIR / "pointwise_comparison_normalized.csv"
df_raw.round({
    "Frequency, Hz": 6, "RMS Voltage Model, V": 6, "RMS Voltage Experiment, V": 6,
    "Output Power Model, mW": 9, "Output Power Experiment, mW": 9,
    "Relative Error RMS, %": 3, "Relative Error Power, %": 3
}).to_csv(csv_raw, index=False, encoding="utf-8-sig")
df_norm.round({
    "Frequency, Hz": 6, "RMS Voltage Model, V (raw)": 6, "RMS Voltage Experiment, V (raw)": 6,
    "RMS Voltage Model (normalized)": 6, "RMS Voltage Experiment (normalized)": 6,
    "Power Model (normalized, V^2)": 6, "Power Experiment (normalized, V^2)": 6,
    "Relative Error RMS, % (normalized)": 3, "Relative Error Power, % (normalized)": 3
}).to_csv(csv_norm, index=False, encoding="utf-8-sig")

# === СОХРАНЕНИЕ В XLSX (два листа) + форматирование
def save_xlsx_two_sheets(raw_df: pd.DataFrame, norm_df: pd.DataFrame, xlsx_path: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.formatting.rule import ColorScaleRule

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        raw_df.to_excel(writer, sheet_name="RAW", index=False)
        norm_df.to_excel(writer, sheet_name="NORMALIZED", index=False)

    wb = load_workbook(xlsx_path)

    def format_sheet(ws, header_to_fmt, error_headers):
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = max_len + 2
        col_map = {cell.value: i+1 for i, cell in enumerate(ws[1])}
        for hdr, fmt in header_to_fmt.items():
            idx = col_map.get(hdr)
            if idx is None: 
                continue
            for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                for cell in row:
                    cell.number_format = fmt
        for hdr in error_headers:
            idx = col_map.get(hdr)
            if idx is None:
                continue
            c1 = ws.cell(row=2, column=idx).coordinate
            c2 = ws.cell(row=ws.max_row, column=idx).coordinate
            ws.conditional_formatting.add(
                f"{c1}:{c2}",
                ColorScaleRule(
                    start_type='min', start_color='63BE7B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='F8696B'
                )
            )

    format_sheet(
        wb["RAW"],
        header_to_fmt={
            "Frequency, Hz": "0.00",
            "RMS Voltage Model, V": "0.000000",
            "RMS Voltage Experiment, V": "0.000000",
            "Output Power Model, mW": "0.000000000",
            "Output Power Experiment, mW": "0.000000000",
            "Relative Error RMS, %": "0.000",
            "Relative Error Power, %": "0.000",
        },
        error_headers=["Relative Error RMS, %", "Relative Error Power, %"],
    )

    format_sheet(
        wb["NORMALIZED"],
        header_to_fmt={
            "Frequency, Hz": "0.00",
            "RMS Voltage Model, V (raw)": "0.000000",
            "RMS Voltage Experiment, V (raw)": "0.000000",
            "RMS Voltage Model (normalized)": "0.000000",
            "RMS Voltage Experiment (normalized)": "0.000000",
            "Power Model (normalized, V^2)": "0.000000",
            "Power Experiment (normalized, V^2)": "0.000000",
            "Relative Error RMS, % (normalized)": "0.000",
            "Relative Error Power, % (normalized)": "0.000",
        },
        error_headers=["Relative Error RMS, % (normalized)", "Relative Error Power, % (normalized)"],
    )

    wb.save(xlsx_path)

xlsx = OUT_DIR / "pointwise_comparison_both.xlsx"
save_xlsx_two_sheets(df_raw.round(9), df_norm.round(9), xlsx)
FIGSIZE_INCH = (6, 3)    # Соотношение 2:1 (ширина:высота)
SAVE_DPI = 1200
FONT_FAMILY = "Arial"     # если нет — возьмётся ближайший sans-serif
FONT_SIZE = 12
XTICK_STEP = 0.5          # Шаг делений по оси X в Гц
plt.rcParams.update({
    "font.family": FONT_FAMILY,
    "font.size": FONT_SIZE,
    "savefig.dpi": SAVE_DPI,
    "axes.labelsize": FONT_SIZE,
    "axes.titlesize": FONT_SIZE,
    "legend.fontsize": FONT_SIZE,
    "xtick.labelsize": FONT_SIZE,
    "ytick.labelsize": FONT_SIZE,
    "axes.formatter.use_mathtext": True,
})
print("[OK] Saved:", csv_raw)
print("[OK] Saved:", csv_norm)
print("[OK] Saved:", xlsx)
print("Done ->", OUT_DIR.resolve())
plot_self_induction_power(
    model_csv=MODEL_CSV,
    out_dir=OUT_DIR,
    r_load_ohm=R_LOAD_OHM,
    range_select=RANGE_SELECT,
    xtick_step=XTICK_STEP,
    figsize_inch=FIGSIZE_INCH,
    filename="self_induction_power1.png",
)
